"""Интеграционные тесты end-to-end для выгрузки курсов и свечей."""

import os
import tempfile
import time
from datetime import date
from unittest.mock import Mock, patch

import openpyxl
import pyarrow.parquet as pq
import pytest

from src.cli.main import main
from src.models.exchange_rate import ExchangeRateRecord
from src.services.cbr_client import CBRClient
from src.services.parquet_writer import ParquetWriter
from src.services.moex_client import MoexClient
from src.utils.date_utils import get_last_7_days


class TestEndToEndFlow:
    """Полный E2E-поток выгрузки и сохранения Parquet."""

    @patch("src.services.cbr_client.requests.Session")
    def test_end_to_end_with_mocked_api(self, mock_session_class):
        """Полный поток от CLI до файла с моканным API ЦБ."""
        # Мокаем XML-ответ
        xml_content = """<?xml version="1.0" encoding="windows-1251"?>
<ValCurs ID="R01235" DateRange1="25.11.2025" DateRange2="01.12.2025" name="Foreign Currency Market Dynamic">
    <Record Date="25.11.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,5000</Value>
        <VunitRate>78,50</VunitRate>
    </Record>
    <Record Date="26.11.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,5500</Value>
        <VunitRate>78,55</VunitRate>
    </Record>
    <Record Date="27.11.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,6000</Value>
        <VunitRate>78,60</VunitRate>
    </Record>
    <Record Date="28.11.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,6500</Value>
        <VunitRate>78,65</VunitRate>
    </Record>
    <Record Date="29.11.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,7000</Value>
        <VunitRate>78,70</VunitRate>
    </Record>
    <Record Date="30.11.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,7500</Value>
        <VunitRate>78,75</VunitRate>
    </Record>
    <Record Date="01.12.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,8000</Value>
        <VunitRate>78,80</VunitRate>
    </Record>
</ValCurs>"""

        encoded_content = xml_content.encode("windows-1251")
        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.content = encoded_content
        mock_session = Mock()
        mock_session.get.return_value = mock_response
        mock_session_class.return_value = mock_session

        with tempfile.TemporaryDirectory() as tmpdir:
        # Переходим во временный каталог
            original_cwd = os.getcwd()
            try:
                os.chdir(tmpdir)

                # Запуск CLI (мок)
                with patch(
                    "src.cli.main.ParquetWriter.write_exchange_rates"
                ) as mock_write:
                    mock_write.return_value = os.path.join(tmpdir, "test.parquet")
                    exit_code = main()

                # Код возврата
                assert exit_code == 0

                # Parquet writer вызван
                assert mock_write.called
                call_args = mock_write.call_args
                records = call_args[0][0]
                metadata = call_args[0][1]

                # Проверяем записи
                assert len(records) == 7
                assert all(isinstance(r, ExchangeRateRecord) for r in records)

                # Проверяем метаданные
                assert "report_date" in metadata
                assert "period_start" in metadata
                assert "period_end" in metadata
                assert metadata["data_source"] == "CBR"
            finally:
                os.chdir(original_cwd)

    @patch("src.services.cbr_client.requests.Session")
    def test_cbr_client_and_parquet_writer_integration(self, mock_session_class):
        """Интеграция клиента ЦБ и ParquetWriter."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Мокаем ответ ЦБ
            xml_content = """<?xml version="1.0" encoding="windows-1251"?>
<ValCurs ID="R01235" DateRange1="25.11.2025" DateRange2="01.12.2025" name="Foreign Currency Market Dynamic">
    <Record Date="25.11.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,5000</Value>
        <VunitRate>78,50</VunitRate>
    </Record>
    <Record Date="26.11.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,5500</Value>
        <VunitRate>78,55</VunitRate>
    </Record>
</ValCurs>"""
            encoded_content = xml_content.encode("windows-1251")
            mock_response = Mock()
            mock_response.status_code = 200
            mock_response.content = encoded_content
            mock_session = Mock()
            mock_session.get.return_value = mock_response
            mock_session_class.return_value = mock_session

            # Получаем данные
            client = CBRClient()
            start_date = date(2025, 11, 25)
            end_date = date(2025, 12, 1)
            records = client.get_exchange_rates(start_date, end_date)

            # Пишем в Parquet
            writer = ParquetWriter()
            metadata = {
                "report_date": "2025-12-02",
                "period_start": "2025-11-25",
                "period_end": "2025-12-01",
                "data_source": "CBR",
            }
            filename = writer.write_exchange_rates(records, metadata, tmpdir)

            # Файл существует и читается
            assert os.path.exists(filename)

            # Читаем и проверяем данные
            table = pq.read_table(filename)
            df = table.to_pandas()
            assert len(df) == 7

            # Проверяем метаданные (контекст для закрытия)
            with pq.ParquetFile(filename) as parquet_file:
                file_metadata = parquet_file.metadata.metadata
                decoded_metadata = {
                    k.decode("utf-8"): v.decode("utf-8")
                    for k, v in file_metadata.items()
                }
                assert decoded_metadata["data_source"] == "CBR"


class TestCLICommandExecution:
    """Сценарии выполнения CLI."""

    @patch("src.services.cbr_client.requests.Session")
    def test_cli_success_scenario(self, mock_session_class):
        """Успешное выполнение CLI."""
        xml_content = """<?xml version="1.0" encoding="windows-1251"?>
<ValCurs ID="R01235" DateRange1="25.11.2025" DateRange2="01.12.2025" name="Foreign Currency Market Dynamic">
    <Record Date="25.11.2025" Id="R01235">
        <Nominal>1</Nominal>
        <Value>78,5000</Value>
        <VunitRate>78,50</VunitRate>
    </Record>
</ValCurs>"""
        encoded_content = xml_content.encode("windows-1251")
        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.content = encoded_content
        mock_session = Mock()
        mock_session.get.return_value = mock_response
        mock_session_class.return_value = mock_session

        with tempfile.TemporaryDirectory() as tmpdir:
            original_cwd = os.getcwd()
            try:
                os.chdir(tmpdir)
                exit_code = main()
                assert exit_code == 0
            finally:
                os.chdir(original_cwd)

    @patch("src.services.cbr_client.requests.Session")
    def test_cli_api_error_scenario(self, mock_session_class):
        """Обработка CLI ошибок API."""
        import requests

        mock_response = Mock()
        mock_response.status_code = 500
        http_error = requests.HTTPError("500 Internal Server Error")
        http_error.response = mock_response
        mock_response.raise_for_status.side_effect = http_error
        mock_session = Mock()
        mock_session.get.return_value = mock_response
        mock_session_class.return_value = mock_session

        exit_code = main()
        assert exit_code == 1  # EXIT_CBR_API_ERROR

    @patch("src.services.cbr_client.requests.Session")
    def test_cli_network_error_scenario(self, mock_session_class):
        """Обработка CLI сетевых ошибок."""
        import requests

        mock_session = Mock()
        mock_session.get.side_effect = requests.Timeout("Request timed out")
        mock_session_class.return_value = mock_session

        exit_code = main()
        assert exit_code == 2  # EXIT_NETWORK_ERROR


class TestMoexEndToEnd:
    """Интеграционные тесты потока moex-lqdt."""

    @patch("src.services.moex_client.requests.Session")
    def test_moex_cli_success_creates_xlsx(self, mock_session_class, tmp_path):
        payload = {
            "candles": {
                "columns": [
                    "open",
                    "close",
                    "high",
                    "low",
                    "value",
                    "volume",
                    "begin",
                    "end",
                    "boardid",
                ],
                "data": [
                    [
                        10.0,
                        11.0,
                        12.0,
                        9.5,
                        0.0,
                        100.0,
                        "2025-11-28 10:00:00",
                        "2025-11-28 18:45:00",
                        "TQTF",
                    ],
                    [
                        11.0,
                        12.0,
                        13.0,
                        10.5,
                        0.0,
                        200.0,
                        "2025-11-30 10:00:00",
                        "2025-11-30 18:45:00",
                        "TQTF",
                    ],
                ],
            }
        }
        mock_response = Mock()
        mock_response.raise_for_status.return_value = None
        mock_response.json.return_value = payload
        mock_session = Mock()
        mock_session.get.return_value = mock_response
        mock_session_class.return_value = mock_session

        original_cwd = os.getcwd()
        try:
            os.chdir(tmp_path)
            with patch("sys.argv", ["python", "moex-lqdt"]):
                exit_code = main()
            assert exit_code == 0
            files = list(tmp_path.glob("lqdt_tqtf_*.xlsx"))
            assert len(files) == 1
            wb = openpyxl.load_workbook(files[0])
            sheet = wb.active
            assert sheet.max_row == 8  # header + 7 rows
            assert sheet.max_column == 6
        finally:
            os.chdir(original_cwd)

    @patch("src.services.moex_client.requests.Session")
    def test_moex_cli_api_error(self, mock_session_class):
        import requests

        mock_response = Mock()
        http_error = requests.HTTPError("503")
        mock_response.raise_for_status.side_effect = http_error
        mock_session = Mock()
        mock_session.get.return_value = mock_response
        mock_session_class.return_value = mock_session

        with patch("sys.argv", ["python", "moex-lqdt"]):
            exit_code = main()
        assert exit_code == 1  # EXIT_API_ERROR

    @patch("src.services.moex_client.requests.Session")
    def test_moex_cli_handles_missing_days(self, mock_session_class, tmp_path):
        payload = {
            "candles": {
                "columns": [
                    "open",
                    "close",
                    "high",
                    "low",
                    "value",
                    "volume",
                    "begin",
                    "end",
                    "boardid",
                ],
                "data": [
                    [
                        10.0,
                        11.0,
                        12.0,
                        9.5,
                        0.0,
                        100.0,
                        "2025-11-28 10:00:00",
                        "2025-11-28 18:45:00",
                        "TQTF",
                    ],
                ],
            }
        }
        mock_response = Mock()
        mock_response.raise_for_status.return_value = None
        mock_response.json.return_value = payload
        mock_session = Mock()
        mock_session.get.return_value = mock_response
        mock_session_class.return_value = mock_session

        original_cwd = os.getcwd()
        try:
            os.chdir(tmp_path)
            with patch("sys.argv", ["python", "moex-lqdt"]):
                exit_code = main()
            assert exit_code == 0
            files = list(tmp_path.glob("lqdt_tqtf_*.xlsx"))
            assert len(files) == 1
            wb = openpyxl.load_workbook(files[0])
            sheet = wb.active
            # Строки на весь период даже при пропусках
            assert sheet.max_row == 8
            # Вторая дата (пропуск) должна иметь пустые значения
            second_row = [cell.value for cell in sheet[3]]
            assert all(value is None for value in second_row[1:])
        finally:
            os.chdir(original_cwd)

    @patch("src.services.moex_client.requests.Session")
    def test_moex_cli_performance_under_20_seconds(
        self, mock_session_class, tmp_path
    ):
        """SC-001: выполнение сценария moex-lqdt занимает <20 секунд при моках."""
        payload = {
            "candles": {
                "columns": [
                    "open",
                    "close",
                    "high",
                    "low",
                    "value",
                    "volume",
                    "begin",
                    "end",
                    "boardid",
                ],
                "data": [
                    [
                        10.0,
                        11.0,
                        12.0,
                        9.5,
                        0.0,
                        100.0,
                        "2025-11-28 10:00:00",
                        "2025-11-28 18:45:00",
                        "TQTF",
                    ],
                ],
            }
        }

        mock_response = Mock()
        mock_response.raise_for_status.return_value = None
        mock_response.json.return_value = payload
        mock_session = Mock()
        mock_session.get.return_value = mock_response
        mock_session_class.return_value = mock_session

        original_cwd = os.getcwd()
        start = time.perf_counter()
        try:
            os.chdir(tmp_path)
            with patch("sys.argv", ["python", "moex-lqdt"]):
                exit_code = main()
        finally:
            os.chdir(original_cwd)

        duration = time.perf_counter() - start
        assert exit_code == 0
        assert duration < 20
        assert list(tmp_path.glob("lqdt_tqtf_*.xlsx"))
