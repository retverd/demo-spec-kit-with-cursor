"""Unit tests for XLSXWriter."""

import os
from datetime import date, timedelta

import openpyxl

from src.models.candles import CandleRecord
from src.services.xlsx_writer import XLSXWriter


def _seven_day_records():
    start = date(2025, 11, 28)
    records = []
    for i in range(7):
        day = start + timedelta(days=i)
        # Make second day empty to test None handling
        if i == 1:
            records.append(
                CandleRecord(
                    date=day,
                    open=None,
                    high=None,
                    low=None,
                    close=None,
                    volume=None,
                )
            )
        else:
            records.append(
                CandleRecord(
                    date=day,
                    open=10.0 + i,
                    high=11.0 + i,
                    low=9.5 + i,
                    close=10.5 + i,
                    volume=100.0 + i,
                )
            )
    return records


def test_write_creates_xlsx_with_headers_and_rows(tmp_path):
    writer = XLSXWriter()
    records = _seven_day_records()
    filename = writer.write_candles(
        records,
        output_dir=tmp_path,
        period_start=records[0].date,
        period_end=records[-1].date,
        report_date=date(2025, 12, 5),
    )
    
    assert os.path.exists(filename)
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    
    headers = [cell.value for cell in sheet[1]]
    assert headers == ["Date", "Open", "High", "Low", "Close", "Volume"]
    # 1 header row + 7 data rows
    assert sheet.max_row == 8
    
    first_data_row = [cell.value for cell in sheet[2]]
    assert first_data_row[0] == records[0].date.isoformat()
    assert first_data_row[1] == records[0].open


def test_write_handles_none_values_as_empty_cells(tmp_path):
    writer = XLSXWriter()
    records = _seven_day_records()
    filename = writer.write_candles(
        records,
        output_dir=tmp_path,
        period_start=records[0].date,
        period_end=records[-1].date,
        report_date=date(2025, 12, 5),
    )
    
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    second_data_row = [cell.value for cell in sheet[3]]  # 3rd row: second record
    # Open/High/Low/Close/Volume should be None -> empty cells
    assert second_data_row[0] == records[1].date.isoformat()
    assert all(value is None for value in second_data_row[1:])


def test_filename_pattern_contains_period_and_report_date(tmp_path):
    writer = XLSXWriter()
    records = _seven_day_records()
    filename = writer.write_candles(
        records,
        output_dir=tmp_path,
        period_start=records[0].date,
        period_end=records[-1].date,
        report_date=date(2025, 12, 5),
    )
    
    basename = os.path.basename(filename)
    assert basename.startswith("lqdt_tqtf_2025-11-28_to_2025-12-04_2025-12-05_")
    assert basename.endswith(".xlsx")

